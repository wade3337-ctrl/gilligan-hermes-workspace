#!/usr/bin/env python3
"""
Bid Output Generator — Takes signals JSON + manual price decisions and builds the deliverable.
Run with: /opt/data/.venv/bin/python bid_output.py --signals signals.json --prices prices.json --output bid.xlsx
"""
import argparse
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_spreadsheet(signals_path, prices_path, output_path, competitor_data=None):
    """Build the final bid spreadsheet from signals and prices."""
    with open(signals_path) as f:
        signals = json.load(f)
    with open(prices_path) as f:
        prices = json.load(f)
    
    city = signals.get('city', 'Unknown')
    tph = signals.get('tph', 130)
    price_buddy = signals.get('price_buddy', {})
    current_rates = signals.get('current_rates', [])
    incumbent = signals.get('incumbent')
    
    wb = openpyxl.Workbook()
    
    # Styles
    hf = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sf = Font(name='Calibri', size=11, bold=True, color='1F4E79')
    sfill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    nf = Font(name='Calibri', size=11)
    bf = Font(name='Calibri', size=11, bold=True)
    crit = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    rev = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    tf = Font(name='Calibri', size=11, bold=True, color='1F4E79')
    tfill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    below = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    above = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin = Border(
        left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7'),
    )
    C = Alignment(horizontal='center', vertical='center')
    L = Alignment(horizontal='left', vertical='center', wrap_text=True)
    R = Alignment(horizontal='right', vertical='center')
    
    # ============ SHEET 1: Cost Proposal ============
    ws = wb.active
    ws.title = "Cost Proposal"
    
    ws.merge_cells('A1:J1')
    ws['A1'] = f'{city.upper()} - MUNICIPAL TREE TRIMMING BID'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
    ws['A1'].alignment = C
    
    ws.merge_cells('A2:J2')
    incumbent_text = f'Incumbent: {incumbent} (CHALLENGER BID)' if incumbent else 'Incumbent: US'
    ws['A2'] = f'Cost Proposal | TPH ${tph}/hr fully-loaded | {incumbent_text}'
    ws['A2'].font = Font(name='Calibri', size=12, bold=True)
    ws['A2'].alignment = C
    
    headers = ['#', 'Line Item', 'Unit', 'Price', 'Current Avg', 'Comp Est', 'Comp Name', 'Grid Floor', 'Margin', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = C; c.border = thin
    
    row = 5
    for i, p in enumerate(prices, 1):
        band = p.get('band', '')
        floor = price_buddy.get(band, {}).get('grid_floor')
        margin = p['price'] - floor if floor else None
        
        ws.cell(row=row, column=1, value=i).font = nf
        ws.cell(row=row, column=1).alignment = C
        ws.cell(row=row, column=2, value=p['desc']).font = nf
        ws.cell(row=row, column=2).alignment = L
        ws.cell(row=row, column=3, value=p['unit']).font = nf
        ws.cell(row=row, column=3).alignment = C
        
        pc = ws.cell(row=row, column=4, value=p['price'])
        pc.font = bf; pc.number_format = '"$"#,##0.00'; pc.alignment = R
        
        # Current rate (find matching)
        cur = p.get('current_rate')
        if cur:
            c = ws.cell(row=row, column=5, value=cur)
            c.number_format = '"$"#,##0.00'; c.alignment = R
        
        # Competitor estimate
        comp = p.get('comp_est')
        if comp:
            c = ws.cell(row=row, column=6, value=comp)
            c.number_format = '"$"#,##0'; c.alignment = R
        
        ws.cell(row=row, column=7, value=p.get('comp_name', '')).font = Font(name='Calibri', size=10, italic=True)
        ws.cell(row=row, column=7).alignment = L
        
        # Floor
        if floor:
            c = ws.cell(row=row, column=8, value=floor)
            c.number_format = '"$"#,##0.00'; c.font = Font(name='Calibri', size=11, color='C00000'); c.alignment = R
            mc = ws.cell(row=row, column=9, value=margin)
            mc.number_format = '"$"+#,##0.00;"$"-#,##0.00'
            mc.font = Font(name='Calibri', size=11, bold=True, color='FF0000' if margin < 0 else '008000')
            mc.alignment = R
        
        ws.cell(row=row, column=10, value=p.get('notes', '')).font = Font(name='Calibri', size=10, italic=True)
        ws.cell(row=row, column=10).alignment = L
        
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin
        
        row += 1
    
    for col, w in {'A':4,'B':34,'C':6,'D':10,'E':11,'F':10,'G':10,'H':10,'I':10,'J':30}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A5'
    
    # ============ SHEET 2: Price Buddy Analysis ============
    if price_buddy:
        ws2 = wb.create_sheet("Price Buddy Analysis")
        ws2['A1'] = f'PRICE BUDDY COST FLOOR - {city.upper()}'
        ws2['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
        ws2.merge_cells('A1:I1'); ws2['A1'].alignment = C
        
        pb_headers = ["Band", "N (jobs)", "Avg $/tree", "Est TPH", "Trim Min", "Hours/tree", "Blended Floor", "Grid Floor", "Our Bid"]
        for col, h in enumerate(pb_headers, 1):
            c = ws2.cell(row=3, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = C; c.border = thin
        
        r = 4
        for band in ["0-6","7-12","13-18","19-24","24-30","31+"]:
            pb = price_buddy.get(band)
            if not pb:
                continue
            our_bid = next((p['price'] for p in prices if p.get('band') == band), None)
            ws2.cell(row=r, column=1, value=band).border = thin
            ws2.cell(row=r, column=2, value=pb['n']).number_format = '#,##0'
            ws2.cell(row=r, column=3, value=pb['avg_price']).number_format = '"$"#,##0.00'
            ws2.cell(row=r, column=4, value=pb['est_tph'])
            ws2.cell(row=r, column=5, value=pb['trim_mins'])
            ws2.cell(row=r, column=6, value=pb['est_hours']).number_format = '0.000'
            ws2.cell(row=r, column=7, value=pb['blended_floor']).number_format = '"$"#,##0.00'
            ws2.cell(row=r, column=8, value=pb['grid_floor']).number_format = '"$"#,##0.00'
            if our_bid:
                ws2.cell(row=r, column=9, value=our_bid).number_format = '"$"#,##0.00'
            
            fill = below if (our_bid and pb['grid_floor'] > our_bid) else (above if our_bid else None)
            if fill:
                for col in range(1, 10):
                    ws2.cell(row=r, column=col).fill = fill
            for col in range(1, 10):
                ws2.cell(row=r, column=col).border = thin
                ws2.cell(row=r, column=col).alignment = R if col > 1 else C
            r += 1
        
        for col, w in {'A':10,'B':10,'C':12,'D':10,'E':10,'F':10,'G':14,'H':14,'I':10}.items():
            ws2.column_dimensions[col].width = w
    
    # ============ SHEET 3: Current Rates Detail ============
    if current_rates:
        ws3 = wb.create_sheet("Current Rates Detail")
        ws3['A1'] = f'OUR CURRENT CONTRACT RATES - {city.upper()}'
        ws3['A1'].font = Font(name='Calibri', size=12, bold=True)
        
        hdrs = ["Line Item", "Size Code", "UOM", "Avg Price", "Min", "Max", "N Projects"]
        for col, h in enumerate(hdrs, 1):
            c = ws3.cell(row=3, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = C; c.border = thin
        
        for i, r in enumerate(current_rates, 4):
            ws3.cell(row=i, column=1, value=r.get('line_item', '')).border = thin
            ws3.cell(row=i, column=2, value=r.get('size_code', '')).border = thin
            ws3.cell(row=i, column=3, value=r.get('uom', '')).border = thin
            ws3.cell(row=i, column=4, value=r.get('avg_price', 0)).number_format = '"$"#,##0.00'
            ws3.cell(row=i, column=5, value=r.get('min', 0)).number_format = '"$"#,##0.00'
            ws3.cell(row=i, column=6, value=r.get('max', 0)).number_format = '"$"#,##0.00'
            ws3.cell(row=i, column=7, value=r.get('n_projects', 0))
            for col in range(1, 8):
                ws3.cell(row=i, column=col).border = thin
                ws3.cell(row=i, column=col).alignment = R if col >= 4 else L
        
        for col, w in {'A':38,'B':12,'C':6,'D':12,'E':10,'F':10,'G':10}.items():
            ws3.column_dimensions[col].width = w
    
    # Save
    wb.save(output_path)
    print(f"Spreadsheet saved: {output_path}")
    return output_path


def send_email(xlsx_path, to_email, subject, body):
    """Send the spreadsheet via email with attachment."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    
    msg = MIMEMultipart()
    msg['From'] = 'bossherman.gsts@gmail.com'
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    
    with open(xlsx_path, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(xlsx_path))
        msg.attach(part)
    
    with open('/opt/data/.secrets/gmail-app.txt') as f:
        password = f.read().strip()
    
    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.login('bossherman.gsts@gmail.com', password)
    server.send_message(msg)
    server.quit()
    print(f"Email sent to {to_email}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Bid Output Generator")
    parser.add_argument("--signals", required=True, help="Path to signals JSON")
    parser.add_argument("--prices", required=True, help="Path to prices JSON")
    parser.add_argument("--output", required=True, help="Output XLSX path")
    args = parser.parse_args()
    
    build_spreadsheet(args.signals, args.prices, args.output)

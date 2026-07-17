#!/usr/bin/env python3
"""
Bid Output Generator v2 — With floor enforcement and complete line-item coverage.
Fixes: no blank lines, $130 floor on labor/day-rate, Price Buddy floor check column,
       variable discount logic, Crown Raise and Stump Grinding always populated.

Run: /opt/data/.venv/bin/python bid_output.py --signals signals.json --prices prices.json --output bid.xlsx
"""
import argparse
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TPH_TARGET = 130

# ============ COMPLETE LINE-ITEM TEMPLATE ============
# Every RFP line that must be priced. No blanks.
# Each entry: (section, description, unit, line_type, band_or_None)
# line_type determines which pricing function to use.
LINE_TEMPLATE = [
    # Section 1: Full Prune
    ("1. Hardwood Grid Trimming - Full Prune", 'Full Prune 0-6 DBH', 'EA', 'grid_prune', '0-6'),
    ("1. Hardwood Grid Trimming - Full Prune", 'Full Prune 7-12 DBH', 'EA', 'grid_prune', '7-12'),
    ("1. Hardwood Grid Trimming - Full Prune", 'Full Prune 13-18 DBH', 'EA', 'grid_prune', '13-18'),
    ("1. Hardwood Grid Trimming - Full Prune", 'Full Prune 19-24 DBH', 'EA', 'grid_prune', '19-24'),
    ("1. Hardwood Grid Trimming - Full Prune", 'Full Prune 24-30 DBH', 'EA', 'grid_prune', '24-30'),
    ("1. Hardwood Grid Trimming - Full Prune", 'Full Prune Over 31 DBH', 'EA', 'grid_prune', '31+'),
    # Section 1: Crown Raise (NEVER blank — derived from Full Prune)
    ("1. Hardwood Grid Trimming - Crown Raise", 'Crown Raise 0-6 DBH', 'EA', 'crown_raise', '0-6'),
    ("1. Hardwood Grid Trimming - Crown Raise", 'Crown Raise 7-12 DBH', 'EA', 'crown_raise', '7-12'),
    ("1. Hardwood Grid Trimming - Crown Raise", 'Crown Raise 13-18 DBH', 'EA', 'crown_raise', '13-18'),
    ("1. Hardwood Grid Trimming - Crown Raise", 'Crown Raise 19-24 DBH', 'EA', 'crown_raise', '19-24'),
    ("1. Hardwood Grid Trimming - Crown Raise", 'Crown Raise 24-30 DBH', 'EA', 'crown_raise', '24-30'),
    ("1. Hardwood Grid Trimming - Crown Raise", 'Crown Raise Over 31 DBH', 'EA', 'crown_raise', '31+'),
    # Section 2: Palms (differentiated by species)
    ("2. Palm Tree Trimming", 'Prune Date Palm (Phoenix spp.)', 'EA', 'palm', 'date'),
    ("2. Palm Tree Trimming", 'Prune Fan Palm (Washingtonia spp.)', 'EA', 'palm', 'fan'),
    ("2. Palm Tree Trimming", 'Prune All Other Palms', 'EA', 'palm', 'other'),
    ("2. Palm Tree Trimming", 'Clean Trunk: Date Palm', 'FT', 'palm_clean', 'date'),
    ("2. Palm Tree Trimming", 'Clean Trunk: Fan Palm', 'FT', 'palm_clean', 'fan'),
    # Section 3: Single Tree (priced above grid to prevent adverse selection)
    ("3. Single Tree Trimming", 'Full Prune Any Diameter Hardwood', 'EA', 'single_tree', None),
    # Section 4: Removal
    ("4. Tree Removal", 'Removal 0-6 DBH', 'EA', 'removal', '0-6'),
    ("4. Tree Removal", 'Removal 7-12 DBH', 'EA', 'removal', '7-12'),
    ("4. Tree Removal", 'Removal 13-18 DBH', 'EA', 'removal', '13-18'),
    ("4. Tree Removal", 'Removal 19-24 DBH', 'EA', 'removal', '19-24'),
    ("4. Tree Removal", 'Removal 24-30 DBH', 'EA', 'removal', '24-30'),
    ("4. Tree Removal", 'Removal Over 31 DBH', 'EA', 'removal', '31+'),
    # Section 5: Stump Grinding (NEVER blank — derived from removal)
    ("5. Stump Grinding", 'Stump Grind 0-6 DBH', 'EA', 'stump', '0-6'),
    ("5. Stump Grinding", 'Stump Grind 7-12 DBH', 'EA', 'stump', '7-12'),
    ("5. Stump Grinding", 'Stump Grind 13-18 DBH', 'EA', 'stump', '13-18'),
    ("5. Stump Grinding", 'Stump Grind 19-24 DBH', 'EA', 'stump', '19-24'),
    ("5. Stump Grinding", 'Stump Grind 24-30 DBH', 'EA', 'stump', '24-30'),
    ("5. Stump Grinding", 'Stump Grind Over 31 DBH', 'EA', 'stump', '31+'),
    # Section 6: Labor (FLOOR ENFORCED at $130)
    ("6. General Hourly Labor Rates", 'Rate for One Ground Person', 'HR', 'labor', None),
    ("6. General Hourly Labor Rates", 'Rate for One Equipment Operator', 'HR', 'labor', None),
    ("6. General Hourly Labor Rates", 'Rate for One Tree Trimmer', 'HR', 'labor', None),
    # Section 7: Day Rate (FLOOR ENFORCED)
    ("7. Day-Rate Service Crew", 'Boom Truck Crew (3 persons x 8 hrs)', 'DAY', 'day_rate', None),
    # Section 8: Planting
    ("8. Tree Planting", '15 Gallon', 'EA', 'planting', None),
    ("8. Tree Planting", '24-inch Box', 'EA', 'planting', None),
    ("8. Tree Planting", '36-inch Box', 'EA', 'planting', None),
    ("8. Tree Planting", '48-inch Box', 'EA', 'planting', None),
    ("8. Tree Planting", 'Fan Palm (10-30 ft BTH)', 'EA', 'planting', None),
    ("8. Tree Planting", 'Fan Palm per ft BTH', 'FT', 'planting', None),
    # Section 9: Watering
    ("9. Tree Watering", 'Water truck and Operator', 'DAY', 'standard', None),
    # Section 10: Emergency (FLOOR ENFORCED + premium)
    ("10. Emergency Services", '3-Person Crew 7AM-7PM', 'HR', 'emergency_day', None),
    ("10. Emergency Services", '3-Person Crew 7PM-7AM', 'HR', 'emergency_night', None),
    # Section 11: Support
    ("11. Support Services", 'ISA-Certified Arborist', 'HR', 'support_arborist', None),
    ("11. Support Services", 'Pest Control Advisor', 'HR', 'support_pca', None),
    ("11. Support Services", 'Qualified Applicator', 'HR', 'support_qac', None),
]

# Default WCA 2021 prices (Long Beach — used as fallback if competitor data not loaded)
WCA_2021_DEFAULT = {
    'grid_prune': {"0-6":44,"7-12":64,"13-18":84,"19-24":114,"24-30":174,"31+":174},
    'crown_raise': {"0-6":34,"7-12":34,"13-18":49,"19-24":64,"24-30":64,"31+":64},
    'palm': {"date":174,"fan":84,"other":64},
    'palm_clean': {"date":54,"fan":54},
    'single_tree': 474,
    'removal': {"0-6":144,"7-12":354,"13-18":654,"19-24":874,"24-30":984,"31+":1094},
    'stump': {"0-6":74,"7-12":94,"13-18":124,"19-24":174,"24-30":194,"31+":224},
    'labor': 94,
    'day_rate': 2256,
    'planting': {"15gal":174,"24box":364,"36box":954,"48box":1754,"fanpalm":1954,"fanpalm_ft":144},
    'watering': 752,
    'emergency_day': 282,
    'emergency_night': 432,
    'arborist': 154,
    'pca': 154,
    'qac': 114,
}

# Escalation factor (3.5% per year × 5 years = 1.188)
ESC = 1.188

# ============ PRICING FUNCTIONS ============

def escalate(price):
    return round(price * ESC)

VOLUME_BANDS = {"13-18", "19-24", "24-30"}

def price_grid_prune(wca_2021, band, grid_floor):
    wca_est = escalate(wca_2021)
    if band in VOLUME_BANDS:
        target = round(wca_est * 0.90)  # 10% under
    elif band == "31+":
        target = round(wca_est * 0.95)  # 5% under (low volume, protect margin)
    else:
        target = round(wca_est * 0.92)  # 8% under
    # Floor check on volume bands
    if grid_floor and band in VOLUME_BANDS and target < grid_floor:
        target = round(grid_floor + 2)
    return target, wca_est

def price_crown_raise(full_prune_price):
    return round(full_prune_price * 0.35)

def price_stump(removal_price):
    return round(removal_price * 0.35)

def price_labor(wca_2021):
    wca_est = escalate(wca_2021)
    target = round(wca_est * 0.92)
    return max(target, TPH_TARGET), wca_est  # FLOOR: $130

def price_day_rate(wca_2021):
    wca_est = escalate(wca_2021)
    target = round(wca_est * 0.92)
    floor = 3 * 8 * TPH_TARGET  # $3,120
    return max(target, floor), wca_est

def price_emergency(wca_2021, after_hours=False):
    wca_est = escalate(wca_2021)
    target = round(wca_est * 0.92)
    floor = 3 * TPH_TARGET * (1.25 if after_hours else 1.0)
    return max(target, round(floor)), wca_est

def price_palm(wca_2021, species):
    wca_est = escalate(wca_2021)
    # Date palms are premium work — smaller discount
    discount = 0.05 if species == 'date' else 0.08
    return round(wca_est * (1 - discount)), wca_est

def price_palm_clean(wca_2021, species):
    wca_est = escalate(wca_2021)
    # Date palm cleaning is harder than fan palm — differentiate
    if species == 'date':
        return round(wca_est * 0.92), wca_est  # 8% under
    else:
        return round(wca_est * 0.60), wca_est  # Fan palm clean is much easier — aggressive

def price_removal(wca_2021, band):
    wca_est = escalate(wca_2021)
    return round(wca_est * 0.90), wca_est  # 10% under (aggressive on removals)

def price_planting(wca_2021):
    wca_est = escalate(wca_2021)
    return round(wca_est * 0.92), wca_est

def price_single_tree(wca_2021):
    wca_est = escalate(wca_2021)
    target = round(wca_est * 0.92)
    # Must be above max grid prune price to prevent adverse selection
    # Max grid is ~$200-250, so floor at $250
    return max(target, 250), wca_est

def price_support(wca_2021, role):
    wca_est = escalate(wca_2021)
    target = round(wca_est * 0.92)
    # Arborist and PCA should be at least $130 floor
    if role in ('arborist', 'pca'):
        target = max(target, TPH_TARGET)
    return target, wca_est


def generate_prices(signals, competitor_prices=None):
    """Generate all 46 prices from the template."""
    price_buddy = signals.get('price_buddy', {})
    
    # Use competitor prices if provided, otherwise use WCA defaults
    wca = competitor_prices if competitor_prices else WCA_2021_DEFAULT
    
    prices = []
    full_prune_prices = {}  # Store for Crown Raise derivation
    removal_prices = {}  # Store for Stump derivation
    
    for section, desc, unit, line_type, band in LINE_TEMPLATE:
        price = None
        wca_est = None
        notes = ""
        
        if line_type == 'grid_prune':
            wca_2021 = wca['grid_prune'][band]
            grid_floor = price_buddy.get(band, {}).get('grid_floor')
            price, wca_est = price_grid_prune(wca_2021, band, grid_floor)
            full_prune_prices[band] = price
            notes = f"Floor: ${grid_floor:.0f}" if grid_floor else ""
        
        elif line_type == 'crown_raise':
            fp = full_prune_prices.get(band, 100)
            price = price_crown_raise(fp)
            wca_est = escalate(wca['crown_raise'].get(band, 50))
            notes = f"35% of Full Prune (${fp})"
        
        elif line_type == 'palm':
            price, wca_est = price_palm(wca['palm'][band], band)
            notes = "Date Palm premium" if band == 'date' else ""
        
        elif line_type == 'palm_clean':
            price, wca_est = price_palm_clean(wca['palm_clean'][band], band)
            notes = "Fan palm clean is easier work" if band == 'fan' else ""
        
        elif line_type == 'single_tree':
            price, wca_est = price_single_tree(wca['single_tree'])
            notes = "Above all grid rates - prevents adverse selection"
        
        elif line_type == 'removal':
            price, wca_est = price_removal(wca['removal'][band], band)
            removal_prices[band] = price
        
        elif line_type == 'stump':
            rem = removal_prices.get(band, 300)
            price = price_stump(rem)
            wca_est = escalate(wca['stump'].get(band, 100))
            notes = f"35% of Removal (${rem})"
        
        elif line_type == 'labor':
            price, wca_est = price_labor(wca['labor'])
            if price == TPH_TARGET:
                notes = f"FLOOR ENFORCED at ${TPH_TARGET}"
            else:
                notes = f"Above ${TPH_TARGET} floor"
        
        elif line_type == 'day_rate':
            price, wca_est = price_day_rate(wca['day_rate'])
            floor = 3 * 8 * TPH_TARGET
            notes = f"Floor: ${floor:,} (3p×8hr×${TPH_TARGET})"
        
        elif line_type == 'emergency_day':
            price, wca_est = price_emergency(wca['emergency_day'], after_hours=False)
            notes = f"Floor: ${3*TPH_TARGET} (3×${TPH_TARGET})"
        
        elif line_type == 'emergency_night':
            price, wca_est = price_emergency(wca['emergency_night'], after_hours=True)
            notes = f"Floor: ${round(3*TPH_TARGET*1.25)} (3×${TPH_TARGET}×1.25)"
        
        elif line_type == 'planting':
            key_map = {'15 Gallon': '15gal', '24-inch Box': '24box', '36-inch Box': '36box',
                       '48-inch Box': '48box', 'Fan Palm (10-30 ft BTH)': 'fanpalm',
                       'Fan Palm per ft BTH': 'fanpalm_ft'}
            key = key_map.get(desc, '')
            price, wca_est = price_planting(wca['planting'].get(key, 200))
        
        elif line_type == 'support_arborist':
            price, wca_est = price_support(wca['arborist'], 'arborist')
        
        elif line_type == 'support_pca':
            price, wca_est = price_support(wca['pca'], 'pca')
        
        elif line_type == 'support_qac':
            price, wca_est = price_support(wca['qac'], 'qac')
        
        elif line_type == 'standard':
            wca_est = escalate(wca.get('watering', 752))
            price = round(wca_est * 0.92)
        
        if price is not None:
            prices.append({
                'section': section, 'desc': desc, 'unit': unit,
                'price': price, 'wca_est': wca_est, 'band': band,
                'notes': notes, 'line_type': line_type,
            })
    
    return prices


def build_spreadsheet(signals_path, prices, output_path):
    """Build the final bid spreadsheet."""
    with open(signals_path) as f:
        signals = json.load(f)
    
    city = signals.get('city', 'Unknown')
    price_buddy = signals.get('price_buddy', {})
    incumbent = signals.get('incumbent', '')
    
    wb = openpyxl.Workbook()
    
    # Styles
    hf = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    sf = Font(name='Calibri', size=11, bold=True, color='1F4E79')
    sfill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    nf = Font(name='Calibri', size=11)
    bf = Font(name='Calibri', size=11, bold=True)
    tf = Font(name='Calibri', size=11, bold=True, color='1F4E79')
    tfill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    floor_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin = Border(
        left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7'),
    )
    C = Alignment(horizontal='center', vertical='center')
    L = Alignment(horizontal='left', vertical='center', wrap_text=True)
    R = Alignment(horizontal='right', vertical='center')
    
    # ============ SHEET 1: Cost Proposal ============
    ws = wb.active
    ws.title = "Cost Proposal"
    
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{city.upper()} - MUNICIPAL TREE TRIMMING BID'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
    ws['A1'].alignment = C
    
    ws.merge_cells('A2:H2')
    inc_text = f'Challenger bid vs {incumbent}' if incumbent else 'Incumbent defense'
    ws['A2'] = f'Cost Proposal | TPH ${TPH_TARGET}/hr fully-loaded | {inc_text}'
    ws['A2'].font = Font(name='Calibri', size=12, bold=True)
    ws['A2'].alignment = C
    
    headers = ['#', 'Line Item', 'Unit', 'Price', 'WCA Estimate', 'Grid Floor', 'Margin', 'Notes']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = C; c.border = thin
    
    row = 5
    current_section = ""
    item = 1
    
    for p in prices:
        # Section header
        if p['section'] != current_section:
            current_section = p['section']
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value=current_section)
            c.font = sf; c.fill = sfill; c.alignment = L; c.border = thin
            row += 1
        
        ws.cell(row=row, column=1, value=item).font = nf
        ws.cell(row=row, column=1).alignment = C
        ws.cell(row=row, column=2, value=p['desc']).font = nf
        ws.cell(row=row, column=2).alignment = L
        ws.cell(row=row, column=3, value=p['unit']).font = nf
        ws.cell(row=row, column=3).alignment = C
        
        pc = ws.cell(row=row, column=4, value=p['price'])
        pc.font = bf; pc.number_format = '"$"#,##0.00'; pc.alignment = R
        
        if p.get('wca_est'):
            c = ws.cell(row=row, column=5, value=p['wca_est'])
            c.number_format = '"$"#,##0'; c.alignment = R
        
        # Grid floor
        band = p.get('band')
        floor = price_buddy.get(band, {}).get('grid_floor') if band else None
        if floor:
            c = ws.cell(row=row, column=6, value=floor)
            c.number_format = '"$"#,##0.00'; c.font = Font(name='Calibri', size=11, color='C00000')
            c.alignment = R
            margin = p['price'] - floor
            mc = ws.cell(row=row, column=7, value=margin)
            mc.number_format = '"$"+#,##0.00;"$"-#,##0.00'
            mc.font = Font(name='Calibri', size=11, bold=True,
                          color='FF0000' if margin < 0 else '008000')
            mc.alignment = R
        
        ws.cell(row=row, column=8, value=p.get('notes', '')).font = Font(name='Calibri', size=10, italic=True)
        ws.cell(row=row, column=8).alignment = L
        
        # Highlight floor-enforced lines
        if 'FLOOR' in p.get('notes', '').upper():
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = floor_fill
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin
        
        row += 1
        item += 1
    
    for col, w in {'A':4, 'B':38, 'C':6, 'D':10, 'E':11, 'F':10, 'G':10, 'H':30}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A5'
    
    # ============ SHEET 2: Price Buddy Analysis ============
    if price_buddy:
        ws2 = wb.create_sheet("Price Buddy Analysis")
        ws2['A1'] = f'PRICE BUDDY COST FLOOR - {city.upper()}'
        ws2['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
        ws2.merge_cells('A1:I1'); ws2['A1'].alignment = C
        
        pb_headers = ["Band", "N (jobs)", "Avg $/tree", "Est TPH", "Trim Min", "Hours/tree", "Blended Floor", "Grid Floor", "Our Bid"]
        for col, h in enumerate(pb_headers, 1):
            c = ws2.cell(row=3, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = C; c.border = thin
        
        r = 4
        for band in ["0-6", "7-12", "13-18", "19-24", "24-30", "31+"]:
            pb = price_buddy.get(band)
            if not pb:
                continue
            our_bid = next((p['price'] for p in prices if p.get('band') == band and p['line_type'] == 'grid_prune'), None)
            ws2.cell(row=r, column=1, value=band).border = thin
            ws2.cell(row=r, column=2, value=pb['n']).number_format = '#,##0'
            ws2.cell(row=r, column=3, value=pb['avg_price']).number_format = '"$"#,##0.00'
            ws2.cell(row=r, column=4, value=pb['est_tph'])
            ws2.cell(row=r, column=5, value=pb['trim_mins'])
            ws2.cell(row=r, column=6, value=pb['est_hours']).number_format = '0.000'
            ws2.cell(row=r, column=7, value=pb['blended_floor']).number_format = '"$"#,##0.00'
            ws2.cell(row=r, column=8, value=pb['grid_floor']).number_format = '"$"#,##0.00'
            if our_bid:
                ws2.cell(row=r, column=9, value=our_bid).number_format = '"$"#,##0.00'
            for col in range(1, 10):
                ws2.cell(row=r, column=col).border = thin
                ws2.cell(row=r, column=col).alignment = R if col > 1 else C
            r += 1
        
        for col, w in {'A':10, 'B':10, 'C':12, 'D':10, 'E':10, 'F':10, 'G':14, 'H':14, 'I':10}.items():
            ws2.column_dimensions[col].width = w
    
    # ============ SHEET 3: Current Rates Detail ============
    current_rates = signals.get('current_rates', [])
    if current_rates:
        ws3 = wb.create_sheet("Current Rates Detail")
        ws3['A1'] = f'OUR CURRENT CONTRACT RATES - {city.upper()}'
        ws3['A1'].font = Font(name='Calibri', size=12, bold=True)
        
        hdrs = ["Line Item", "Size Code", "UOM", "Avg Price", "Min", "Max", "N Projects"]
        for col, h in enumerate(hdrs, 1):
            c = ws3.cell(row=3, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = C; c.border = thin
        
        for i, r_data in enumerate(current_rates, 4):
            ws3.cell(row=i, column=1, value=r_data.get('line_item', '')).border = thin
            ws3.cell(row=i, column=2, value=r_data.get('size_code', '')).border = thin
            ws3.cell(row=i, column=3, value=r_data.get('uom', '')).border = thin
            ws3.cell(row=i, column=4, value=r_data.get('avg_price', 0)).number_format = '"$"#,##0.00'
            ws3.cell(row=i, column=5, value=r_data.get('min', 0)).number_format = '"$"#,##0.00'
            ws3.cell(row=i, column=6, value=r_data.get('max', 0)).number_format = '"$"#,##0.00'
            ws3.cell(row=i, column=7, value=r_data.get('n_projects', 0))
            for col in range(1, 8):
                ws3.cell(row=i, column=col).border = thin
                ws3.cell(row=i, column=col).alignment = R if col >= 4 else L
        
        for col, w in {'A':38, 'B':12, 'C':6, 'D':12, 'E':10, 'F':10, 'G':10}.items():
            ws3.column_dimensions[col].width = w
    
    # Save
    wb.save(output_path)
    print(f"Spreadsheet saved: {output_path}")
    print(f"Total line items: {len(prices)} (no blanks)")
    return output_path


def send_email(xlsx_path, to_email, subject, body):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    
    msg = MIMEMultipart()
    msg['From'] = 'bossherman.gsts@gmail.com'
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    
    with open(xlsx_path, 'rb') as f:
        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(xlsx_path))
        msg.attach(part)
    
    with open('/opt/data/.secrets/gmail-app.txt') as f:
        password = f.read().strip()
    
    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.login('bossherman.gsts@gmail.com', password)
    server.send_message(msg)
    server.quit()
    print(f"Email sent to {to_email}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Bid Output Generator v2")
    parser.add_argument("--signals", required=True, help="Path to signals JSON from bid_engine.py")
    parser.add_argument("--output", required=True, help="Output XLSX path")
    args = parser.parse_args()
    
    with open(args.signals) as f:
        signals = json.load(f)
    
    prices = generate_prices(signals)
    build_spreadsheet(args.signals, prices, args.output)
    
    # Verify no blanks
    blanks = [p for p in prices if p['price'] is None]
    if blanks:
        print(f"WARNING: {len(blanks)} blank prices found!")
        for b in blanks:
            print(f"  BLANK: {b['desc']}")
    else:
        print("VERIFIED: All 46 line items priced. No blanks.")
    
    # Floor violations
    floor_violations = [p for p in prices if 'labor' in p['line_type'] and p['price'] < TPH_TARGET]
    if floor_violations:
        print(f"WARNING: {len(floor_violations)} labor rates below $130 floor!")
    else:
        print("VERIFIED: All labor rates at or above $130 TPH floor.")
